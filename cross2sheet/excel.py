"""
Converts grids to Excel format.
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.colors import Color
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
import collections
from cross2sheet.grid_features import *

class _CellStyle:

    border_names = { 'L' : 'left', 'R' : 'right', 'T' : 'top', 'B' : 'bottom' }

    def __init__(self):
        self.color = None
        self.borders = set()

    def set_style(self,cell):
        if self.color is not None:
            cell.fill=PatternFill(patternType='solid',fgColor=Color('FF%06x'%self.color))
        if self.borders:
            kwa = { self.border_names[b] : Side(style='thick') for b in self.borders }
            cell.border=Border(**kwa)

def write_sheet(grid,ws,text_in_cells=True,text_in_comments=False,leave_white_blank=True):
    styles = collections.defaultdict(_CellStyle)
    for r,c,elt in grid.features:
        cell = ws.cell(row=r+1,column=c+1)
        if isinstance(elt,BackgroundElt):
            if not (elt.color==0xFFFFFF and leave_white_blank):
                styles[r,c].color=elt.color
        elif isinstance(elt,TextElt):
            if text_in_cells:
                cell.value=elt.text
            if text_in_comments:
                cell.comment=Comment(elt.text,'')
        elif isinstance(elt,BorderElt):
            styles[r,c].borders.update(elt.dirs)
    for (r,c),s in styles.items():
        s.set_style(ws.cell(row=r+1,column=c+1))
    for c in range(grid.width):
        ws.column_dimensions[get_column_letter(c+1)].width=3
    # Google Sheets seems to truncate sheets with no data at 10 columns, so
    # make sure the bottom right cell isn't empty
    bottom_right=ws.cell(row=grid.height,column=grid.width)
    if not bottom_right.value:
        bottom_right.value=' '

def to_openpyxl(grid,**kwargs):
    wb = Workbook()
    write_sheet(grid,wb.active,**kwargs)
    return wb

def to_openpyxl_multi(grids,**kwargs):
    wb = Workbook()
    wb.remove_sheet(wb.active)
    for grid in grids:
        ws = wb.create_sheet()
        write_sheet(grid,ws,**kwargs)
    return wb

def save_xlsx(grid,filename,**kwargs):
    return to_openpyxl(grid,**kwargs).save(filename)
